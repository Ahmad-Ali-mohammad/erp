from datetime import date, datetime
from io import BytesIO
from decimal import Decimal

from django.db import transaction
from django.db.models import Sum
from django.http import HttpResponse
from django.utils import timezone
from django.utils.dateparse import parse_date
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from rest_framework import mixins, viewsets
from rest_framework.decorators import action
from rest_framework.exceptions import PermissionDenied, ValidationError
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated

from core.access import (
    ActionBasedRolePermission,
    ROLE_ACCOUNTANT,
    ROLE_ADMIN,
    ROLE_PROJECT_MANAGER,
    ROLE_UNASSIGNED,
    RowLevelScopeMixin,
)
from core.audit import AuditLogMixin
from projects.models import Project, ProjectCostRecord
from projects.services import settle_project_cost_records_by_source, sync_project_cost_records_by_source
from .models import (
    Account,
    BankAccount,
    BankReconciliationSession,
    BankStatement,
    ExchangeRate,
    FiscalPeriod,
    Invoice,
    InvoiceItem,
    JournalEntry,
    JournalLine,
    Payment,
    PostingRule,
    ProgressBilling,
    PrintSettings,
    RecurringEntryTemplate,
    RevenueRecognitionEntry,
)
from .serializers import (
    AccountSerializer,
    BankAccountSerializer,
    BankReconciliationSessionSerializer,
    BankStatementSerializer,
    ExchangeRateSerializer,
    FiscalPeriodSerializer,
    InvoiceSerializer,
    JournalEntrySerializer,
    PaymentSerializer,
    PostingRuleSerializer,
    ProgressBillingSerializer,
    PrintSettingsSerializer,
    RecurringEntryTemplateSerializer,
    RevenueRecognitionEntrySerializer,
)
from .services.posting_engine import PostingEngine
from .services.printing import get_print_settings, next_invoice_number
from .services.reporting import (
    build_balance_sheet,
    build_general_journal,
    build_general_ledger,
    build_income_statement,
    build_trial_balance,
)
from payments.serializers import PaymentAllocationDetailSerializer
from payments.models import PaymentAllocation

FINANCE_APPROVER_ROLE_SLUGS = {"admin", "accountant"}
FINANCE_READ_ROLES = {ROLE_ADMIN, ROLE_ACCOUNTANT, ROLE_PROJECT_MANAGER, ROLE_UNASSIGNED}
FINANCE_WRITE_ROLES = {ROLE_ADMIN, ROLE_ACCOUNTANT, ROLE_PROJECT_MANAGER, ROLE_UNASSIGNED}
FINANCE_APPROVER_ROLES = {ROLE_ADMIN, ROLE_ACCOUNTANT}
FINANCE_SETUP_ROLES = {ROLE_ADMIN, ROLE_ACCOUNTANT}
LOCKED_PROJECT_STATUSES = {Project.Status.COMPLETED, Project.Status.CANCELLED}


def _is_finance_approver(user) -> bool:
    if not user.is_authenticated:
        return False
    if user.is_superuser:
        return True
    role = getattr(user, "role", None)
    return bool(role and role.slug in FINANCE_APPROVER_ROLE_SLUGS)


def _sync_invoice_payment_status(invoice: Invoice) -> None:
    if invoice.status not in {
        Invoice.InvoiceStatus.ISSUED,
        Invoice.InvoiceStatus.PARTIALLY_PAID,
        Invoice.InvoiceStatus.PAID,
    }:
        return

    confirmed_total = (
        invoice.payments.filter(status=Payment.Status.CONFIRMED).aggregate(total=Sum("amount"))["total"]
        or Decimal("0.00")
    )
    if confirmed_total >= invoice.total_amount and invoice.total_amount > Decimal("0.00"):
        new_status = Invoice.InvoiceStatus.PAID
    elif confirmed_total > Decimal("0.00"):
        new_status = Invoice.InvoiceStatus.PARTIALLY_PAID
    else:
        new_status = Invoice.InvoiceStatus.ISSUED

    if invoice.status != new_status:
        invoice.status = new_status
        invoice.save(update_fields=["status", "updated_at"])


def _assert_project_open(project):
    if project and project.status in LOCKED_PROJECT_STATUSES:
        raise ValidationError({"project": "This project is closed and cannot be modified."})


def _build_invoice_cost_amount_map(invoice: Invoice) -> dict:
    if invoice.cost_code:
        return {invoice.cost_code: invoice.total_amount}

    grouped_map = {}
    for item in invoice.items.select_related("cost_code").all():
        if not item.cost_code:
            continue
        line_subtotal = item.quantity * item.unit_price
        line_tax = line_subtotal * (item.tax_rate / Decimal("100"))
        line_total = line_subtotal + line_tax

        key = item.cost_code_id
        if key in grouped_map:
            grouped_map[key]["amount"] += line_total
        else:
            grouped_map[key] = {"cost_code": item.cost_code, "amount": line_total}

    return {entry["cost_code"]: entry["amount"] for entry in grouped_map.values()}


def _quantize_money(amount: Decimal) -> Decimal:
    return amount.quantize(Decimal("0.01"))


def _approved_progress_billing_subtotal(project, exclude_id=None) -> Decimal:
    queryset = ProgressBilling.objects.filter(
        project=project,
        status__in=[ProgressBilling.Status.APPROVED, ProgressBilling.Status.INVOICED],
    )
    if exclude_id:
        queryset = queryset.exclude(id=exclude_id)
    return queryset.aggregate(total=Sum("subtotal"))["total"] or Decimal("0.00")


def _approved_revenue_total(project, exclude_id=None) -> Decimal:
    queryset = RevenueRecognitionEntry.objects.filter(
        project=project,
        status=RevenueRecognitionEntry.Status.APPROVED,
    )
    if exclude_id:
        queryset = queryset.exclude(id=exclude_id)
    return queryset.aggregate(total=Sum("recognized_amount"))["total"] or Decimal("0.00")


def _next_progress_invoice_number(billing_number: str) -> str:
    base_number = f"PB-{billing_number}"
    candidate = base_number
    counter = 1

    while Invoice.objects.filter(invoice_number=candidate).exists():
        candidate = f"{base_number}-{counter:02d}"
        counter += 1
    return candidate


def _next_related_entry_number(prefix: str, base_entry_number: str) -> str:
    candidate = f"{prefix}-{base_entry_number}"
    counter = 1
    while JournalEntry.objects.filter(entry_number=candidate).exists():
        candidate = f"{prefix}-{base_entry_number}-{counter:02d}"
        counter += 1
    return candidate


def _parse_date_param(value: str | None, *, field_name: str) -> date | None:
    if not value:
        return None
    parsed = parse_date(value)
    if not parsed:
        raise ValidationError({field_name: "Invalid date format. Use YYYY-MM-DD."})
    return parsed


def _parse_excel_date(value, *, field_name: str, row_number: int, errors: list[dict]) -> date | None:
    if value is None or value == "":
        errors.append({"row": row_number, "field": field_name, "message": "Date is required."})
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    parsed = parse_date(str(value))
    if not parsed:
        errors.append({"row": row_number, "field": field_name, "message": "Invalid date format. Use YYYY-MM-DD."})
        return None
    return parsed


def _parse_decimal(
    value,
    *,
    field_name: str,
    row_number: int,
    errors: list[dict],
    quantize: Decimal | None = Decimal("0.01"),
) -> Decimal:
    if value is None or value == "":
        return Decimal("0.00")
    try:
        parsed = Decimal(str(value))
        if quantize is not None:
            parsed = parsed.quantize(quantize)
        return parsed
    except Exception:
        errors.append({"row": row_number, "field": field_name, "message": "Invalid numeric value."})
        return Decimal("0.00")


def _next_journal_entry_number(entry_date: date | None = None) -> str:
    effective_date = entry_date or timezone.localdate()
    base = f"JE-{effective_date.strftime('%Y%m%d')}-"
    counter = 1
    while JournalEntry.objects.filter(entry_number=f"{base}{counter:03d}").exists():
        counter += 1
    return f"{base}{counter:03d}"


def _extract_validation_text(exc: ValidationError) -> str:
    detail = getattr(exc, "detail", None)
    if isinstance(detail, dict):
        for value in detail.values():
            if isinstance(value, list) and value:
                return str(value[0])
            return str(value)
    if isinstance(detail, list) and detail:
        return str(detail[0])
    return str(exc)


def _normalize_source_currency(source_object) -> None:
    currency = getattr(source_object, "currency", None)
    if not currency and hasattr(source_object, "invoice"):
        currency = getattr(source_object.invoice, "currency", None)
    if not currency and hasattr(source_object, "project"):
        currency = getattr(source_object.project, "currency", None)

    normalized = str(currency or "KWD").upper()
    if normalized not in {"KWD", "USD"}:
        normalized = "KWD"
    setattr(source_object, "currency", normalized)


def _auto_post_operational_event(
    *,
    source_module: str,
    source_event: str,
    source_object,
    entry_date: date,
    description: str,
    posted_by,
    idempotency_key: str,
):
    _normalize_source_currency(source_object)
    try:
        return PostingEngine.post_from_operational_event(
            source_module=source_module,
            source_event=source_event,
            source_object=source_object,
            entry_date=entry_date,
            description=description,
            posted_by=posted_by,
            idempotency_key=idempotency_key,
            entry_class=JournalEntry.EntryClass.OPERATIONAL_AUTO,
        )
    except ValidationError as exc:
        error_text = _extract_validation_text(exc)
        # Keep backwards compatibility when posting rules are not configured yet.
        if "No active posting rule configured" in error_text:
            return None
        raise


def _ensure_journal_entry_balanced(entry: JournalEntry) -> None:
    totals = entry.lines.aggregate(total_debit=Sum("debit"), total_credit=Sum("credit"))
    debit_total = totals["total_debit"] or Decimal("0.00")
    credit_total = totals["total_credit"] or Decimal("0.00")
    if debit_total != credit_total:
        raise ValidationError({"lines": "Journal entry must be balanced before posting."})


class AccountViewSet(AuditLogMixin, viewsets.ModelViewSet):
    queryset = Account.objects.select_related("parent").all()
    serializer_class = AccountSerializer
    permission_classes = [ActionBasedRolePermission]
    action_role_map = {
        "list": FINANCE_READ_ROLES,
        "retrieve": FINANCE_READ_ROLES,
        "create": FINANCE_SETUP_ROLES,
        "update": FINANCE_SETUP_ROLES,
        "partial_update": FINANCE_SETUP_ROLES,
        "destroy": {ROLE_ADMIN},
    }
    filterset_fields = ["account_type", "is_active", "is_control_account", "parent"]
    search_fields = ["code", "name"]
    ordering_fields = ["code", "name", "created_at"]


class JournalEntryViewSet(AuditLogMixin, RowLevelScopeMixin, viewsets.ModelViewSet):
    queryset = JournalEntry.objects.prefetch_related("lines").select_related(
        "project",
        "created_by",
        "posted_by",
        "period",
        "reversal_of",
        "correction_root",
    )
    serializer_class = JournalEntrySerializer
    permission_classes = [ActionBasedRolePermission]
    user_scope_fields = ("created_by", "project__created_by")
    action_role_map = {
        "list": FINANCE_READ_ROLES,
        "retrieve": FINANCE_READ_ROLES,
        "create": FINANCE_WRITE_ROLES,
        "update": FINANCE_WRITE_ROLES,
        "partial_update": FINANCE_WRITE_ROLES,
        "destroy": {ROLE_ADMIN, ROLE_ACCOUNTANT},
        "post": FINANCE_APPROVER_ROLES,
        "reverse": FINANCE_APPROVER_ROLES,
        "correct": FINANCE_APPROVER_ROLES,
        "export": FINANCE_READ_ROLES,
        "import_entries": FINANCE_WRITE_ROLES,
        "import_template": FINANCE_READ_ROLES,
    }
    filterset_fields = [
        "status",
        "entry_class",
        "project",
        "period",
        "entry_date",
        "source_module",
        "source_event",
    ]
    search_fields = ["entry_number", "description", "project__code", "source_module", "source_event"]
    ordering_fields = ["entry_date", "created_at", "entry_number", "posted_at"]

    def get_queryset(self):
        return self.apply_row_level_scope(super().get_queryset())

    def perform_create(self, serializer):
        instance = serializer.save(created_by=self.request.user)
        self.log_action(action="create", instance=instance, changes=serializer.validated_data)

    def perform_update(self, serializer):
        if serializer.instance.status != JournalEntry.Status.DRAFT:
            raise ValidationError({"status": "Only draft journal entries can be modified."})
        changes = self._build_changes(serializer.instance, serializer.validated_data)
        instance = serializer.save()
        self.log_action(action="update", instance=instance, changes=changes)

    def perform_destroy(self, instance):
        if instance.status != JournalEntry.Status.DRAFT:
            raise ValidationError({"status": "Only draft journal entries can be deleted."})
        self.log_action(action="delete", instance=instance)
        instance.delete()

    @action(detail=True, methods=["post"])
    def post(self, request, pk=None):
        entry = self.get_object()
        if entry.status != JournalEntry.Status.DRAFT:
            raise ValidationError({"status": "Only draft entries can be posted."})
        if entry.period and entry.period.status == FiscalPeriod.Status.HARD_CLOSED:
            raise ValidationError({"period": "Cannot post entries into a hard-closed period."})

        _ensure_journal_entry_balanced(entry)

        entry.status = JournalEntry.Status.POSTED
        entry.posted_at = timezone.now()
        entry.posted_by = request.user
        entry.save(update_fields=["status", "posted_at", "posted_by", "updated_at"])
        return Response(self.get_serializer(entry).data)

    @action(detail=True, methods=["post"])
    def reverse(self, request, pk=None):
        entry = self.get_object()
        if entry.status != JournalEntry.Status.POSTED:
            raise ValidationError({"status": "Only posted entries can be reversed."})
        if entry.reversals.filter(status=JournalEntry.Status.POSTED).exists():
            raise ValidationError({"status": "This entry already has a posted reversal."})

        with transaction.atomic():
            reversal = JournalEntry.objects.create(
                entry_number=_next_related_entry_number("REV", entry.entry_number),
                entry_date=timezone.localdate(),
                description=f"Reversal of {entry.entry_number}",
                status=JournalEntry.Status.POSTED,
                entry_class=JournalEntry.EntryClass.CORRECTION,
                currency=entry.currency,
                fx_rate_to_base=entry.fx_rate_to_base,
                period=entry.period,
                posted_at=timezone.now(),
                posted_by=request.user,
                reversal_of=entry,
                correction_root=entry.correction_root or entry,
                project=entry.project,
                created_by=request.user,
            )
            for line in entry.lines.all():
                JournalLine.objects.create(
                    entry=reversal,
                    account=line.account,
                    description=f"Reversal - {line.description}".strip(" -"),
                    debit=line.credit,
                    credit=line.debit,
                    debit_foreign=line.credit_foreign,
                    credit_foreign=line.debit_foreign,
                    project=line.project or entry.project,
                    cost_center_code=line.cost_center_code,
                )
            entry.status = JournalEntry.Status.REVERSED
            entry.save(update_fields=["status", "updated_at"])

        return Response(
            {
                "entry": self.get_serializer(entry).data,
                "reversal_entry": self.get_serializer(reversal).data,
            }
        )

    @action(detail=True, methods=["post"])
    def correct(self, request, pk=None):
        entry = self.get_object()
        if entry.status not in {JournalEntry.Status.POSTED, JournalEntry.Status.REVERSED}:
            raise ValidationError({"status": "Only posted/reversed entries can be corrected."})

        reason = str(request.data.get("reason", "")).strip()
        correction_lines = request.data.get("lines")
        if not isinstance(correction_lines, list) or len(correction_lines) == 0:
            raise ValidationError({"lines": "Correction requires at least one journal line."})

        debit_total = Decimal("0.00")
        credit_total = Decimal("0.00")
        for line in correction_lines:
            debit_total += Decimal(line.get("debit", "0") or "0")
            credit_total += Decimal(line.get("credit", "0") or "0")
        if debit_total != credit_total:
            raise ValidationError({"lines": "Correction entry must be balanced (debit == credit)."})

        with transaction.atomic():
            if entry.status == JournalEntry.Status.POSTED:
                reverse_response = self.reverse(request, pk=pk)
                reversal_payload = reverse_response.data.get("reversal_entry")
            else:
                reversal_payload = None

            correction_entry = JournalEntry.objects.create(
                entry_number=_next_related_entry_number("COR", entry.entry_number),
                entry_date=timezone.localdate(),
                description=request.data.get("description") or f"Correction of {entry.entry_number}",
                status=JournalEntry.Status.POSTED,
                entry_class=JournalEntry.EntryClass.CORRECTION,
                currency=entry.currency,
                fx_rate_to_base=entry.fx_rate_to_base,
                period=entry.period,
                posted_at=timezone.now(),
                posted_by=request.user,
                correction_root=entry.correction_root or entry,
                project=entry.project,
                created_by=request.user,
            )
            for line in correction_lines:
                JournalLine.objects.create(
                    entry=correction_entry,
                    account_id=line.get("account"),
                    description=str(line.get("description", "")),
                    debit=Decimal(line.get("debit", "0") or "0"),
                    credit=Decimal(line.get("credit", "0") or "0"),
                    debit_foreign=Decimal(line["debit_foreign"]) if line.get("debit_foreign") not in [None, ""] else None,
                    credit_foreign=Decimal(line["credit_foreign"]) if line.get("credit_foreign") not in [None, ""] else None,
                    project_id=line.get("project") or entry.project_id,
                    cost_center_code=str(line.get("cost_center_code", "")),
                )

            correction_detail = getattr(correction_entry, "correction_detail", None)
            if correction_detail:
                correction_detail.reason = reason
                correction_detail.corrected_entry = entry
                correction_detail.save(update_fields=["reason", "corrected_entry", "updated_at"])

        return Response(
            {
                "corrected_entry": self.get_serializer(correction_entry).data,
                "reversal_entry": reversal_payload,
            }
        )

    @action(detail=False, methods=["get"], url_path="export")
    def export(self, request):
        queryset = self.filter_queryset(self.get_queryset()).prefetch_related("lines__account")

        start_date = _parse_date_param(request.query_params.get("start_date"), field_name="start_date")
        end_date = _parse_date_param(request.query_params.get("end_date"), field_name="end_date")
        if start_date and end_date and start_date > end_date:
            raise ValidationError({"start_date": "start_date must be before or equal to end_date."})
        if start_date:
            queryset = queryset.filter(entry_date__gte=start_date)
        if end_date:
            queryset = queryset.filter(entry_date__lte=end_date)

        project_id = request.query_params.get("project")
        if project_id:
            try:
                queryset = queryset.filter(project_id=int(project_id))
            except (TypeError, ValueError):
                raise ValidationError({"project": "project must be an integer id."})

        headers = [
            "entry_number",
            "entry_date",
            "entry_class",
            "description",
            "currency",
            "fx_rate_to_base",
            "project_id",
            "account_code",
            "account_id",
            "debit",
            "credit",
            "line_description",
        ]

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Journal Entries"

        title_text = "Journal Entries Export"
        sheet["A1"] = title_text
        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        sheet["A1"].font = Font(size=14, bold=True)
        sheet["A1"].alignment = Alignment(horizontal="center")

        sheet["A2"] = f"Generated at: {timezone.now().strftime('%Y-%m-%d %H:%M')}"
        sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
        sheet["A2"].alignment = Alignment(horizontal="center")

        header_row_index = 3
        sheet.append(headers)
        header_fill = PatternFill("solid", fgColor="F1F5F9")
        for col_idx in range(1, len(headers) + 1):
            cell = sheet.cell(row=header_row_index, column=col_idx)
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            sheet.column_dimensions[cell.column_letter].width = 18

        for entry in queryset.order_by("entry_date", "entry_number", "id"):
            for line in entry.lines.all():
                sheet.append(
                    [
                        entry.entry_number,
                        entry.entry_date.isoformat(),
                        entry.entry_class,
                        entry.description,
                        entry.currency,
                        str(entry.fx_rate_to_base),
                        entry.project_id or "",
                        line.account.code,
                        line.account_id,
                        str(line.debit),
                        str(line.credit),
                        line.description,
                    ]
                )

        output = BytesIO()
        workbook.save(output)
        output.seek(0)

        filename = f"journal-entries-{timezone.localdate().strftime('%Y%m%d')}.xlsx"
        response = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response

    @action(detail=False, methods=["get"], url_path="import-template")
    def import_template(self, request):
        headers = [
            "entry_number",
            "entry_date",
            "entry_class",
            "description",
            "currency",
            "fx_rate_to_base",
            "project_id",
            "account_code",
            "account_id",
            "debit",
            "credit",
            "line_description",
        ]

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Journal Entries"
        sheet.append(headers)
        header_fill = PatternFill("solid", fgColor="F1F5F9")
        for col_idx in range(1, len(headers) + 1):
            cell = sheet.cell(row=1, column=col_idx)
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            sheet.column_dimensions[cell.column_letter].width = 18

        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        filename = "journal-entries-template.xlsx"
        response = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response

    @action(detail=False, methods=["post"], url_path="import")
    def import_entries(self, request):
        upload = request.FILES.get("file") or request.FILES.get("xlsx")
        if not upload:
            raise ValidationError({"file": "Excel file (.xlsx) is required."})

        try:
            workbook = load_workbook(upload, data_only=True)
        except Exception:
            raise ValidationError({"file": "Unable to read the Excel file."})

        sheet = workbook.active
        header_row_index = None
        header_map: dict[str, int] = {}

        for row_index, row in enumerate(sheet.iter_rows(min_row=1, max_row=10, values_only=True), start=1):
            normalized = [str(cell).strip().lower() if cell is not None else "" for cell in row]
            if "entry_date" in normalized and ("debit" in normalized or "credit" in normalized):
                header_row_index = row_index
                header_map = {name: idx for idx, name in enumerate(normalized) if name}
                break

        if header_row_index is None:
            raise ValidationError({"file": "Header row not found. Ensure the template headers are present."})

        required_headers = {"entry_date", "debit", "credit"}
        missing_headers = [header for header in required_headers if header not in header_map]
        if missing_headers:
            raise ValidationError({"file": f"Missing required columns: {', '.join(missing_headers)}"})

        entries: dict[str, dict] = {}
        errors: list[dict] = []

        for row_number, row in enumerate(sheet.iter_rows(min_row=header_row_index + 1, values_only=True), start=header_row_index + 1):
            if not row or all(cell is None or str(cell).strip() == "" for cell in row):
                continue

            def cell_value(name: str):
                idx = header_map.get(name)
                if idx is None:
                    return None
                if idx >= len(row):
                    return None
                return row[idx]

            entry_date = _parse_excel_date(cell_value("entry_date"), field_name="entry_date", row_number=row_number, errors=errors)
            if entry_date is None:
                continue

            raw_entry_number = str(cell_value("entry_number") or "").strip()
            entry_number = raw_entry_number or _next_journal_entry_number(entry_date)

            entry_class = str(cell_value("entry_class") or "manual").strip() or "manual"
            if entry_class not in {choice for choice, _ in JournalEntry.EntryClass.choices}:
                errors.append(
                    {
                        "row": row_number,
                        "field": "entry_class",
                        "message": f"Invalid entry_class '{entry_class}'.",
                    }
                )
                continue

            description = str(cell_value("description") or "").strip()
            currency = str(cell_value("currency") or "KWD").strip().upper() or "KWD"
            fx_rate = cell_value("fx_rate_to_base")
            fx_rate_value = _parse_decimal(
                fx_rate,
                field_name="fx_rate_to_base",
                row_number=row_number,
                errors=errors,
                quantize=Decimal("0.00000001"),
            )
            if fx_rate_value <= Decimal("0.00"):
                fx_rate_value = Decimal("1.00000000")

            project_id_value = cell_value("project_id")
            project_id = None
            if project_id_value not in (None, ""):
                try:
                    project_id = int(project_id_value)
                except (TypeError, ValueError):
                    errors.append({"row": row_number, "field": "project_id", "message": "project_id must be an integer."})
                    continue

            account_code = str(cell_value("account_code") or "").strip()
            account_id_value = cell_value("account_id")
            account_id = None
            if account_id_value not in (None, ""):
                try:
                    account_id = int(account_id_value)
                except (TypeError, ValueError):
                    errors.append({"row": row_number, "field": "account_id", "message": "account_id must be an integer."})
                    continue

            if not account_id and account_code:
                account_id = Account.objects.filter(code=account_code).values_list("id", flat=True).first()

            if not account_id:
                errors.append(
                    {
                        "row": row_number,
                        "field": "account_code",
                        "message": "Account not found. Provide a valid account_code or account_id.",
                    }
                )
                continue

            debit = _parse_decimal(cell_value("debit"), field_name="debit", row_number=row_number, errors=errors)
            credit = _parse_decimal(cell_value("credit"), field_name="credit", row_number=row_number, errors=errors)
            if (debit > 0 and credit > 0) or (debit <= 0 and credit <= 0):
                errors.append(
                    {
                        "row": row_number,
                        "field": "debit/credit",
                        "message": "Line must contain a debit or credit amount (not both).",
                    }
                )
                continue

            line_description = str(cell_value("line_description") or "").strip()

            entry_payload = entries.setdefault(
                entry_number,
                {
                    "entry_number": entry_number,
                    "entry_date": entry_date.isoformat(),
                    "entry_class": entry_class,
                    "description": description,
                    "currency": currency,
                    "fx_rate_to_base": str(fx_rate_value),
                    "project": project_id,
                    "status": JournalEntry.Status.DRAFT,
                    "lines": [],
                },
            )

            if entry_payload["entry_date"] != entry_date.isoformat():
                errors.append(
                    {
                        "row": row_number,
                        "field": "entry_date",
                        "message": "entry_date must be consistent for the same entry_number.",
                    }
                )
                continue

            entry_payload["lines"].append(
                {
                    "account": account_id,
                    "description": line_description,
                    "debit": str(debit),
                    "credit": str(credit),
                    "project": project_id,
                }
            )

        for entry_number, entry_payload in entries.items():
            debit_total = Decimal("0.00")
            credit_total = Decimal("0.00")
            for line in entry_payload["lines"]:
                debit_total += Decimal(line.get("debit", "0") or "0")
                credit_total += Decimal(line.get("credit", "0") or "0")
            if debit_total != credit_total:
                errors.append(
                    {
                        "entry_number": entry_number,
                        "field": "lines",
                        "message": "Entry is not balanced (debit != credit).",
                    }
                )

        if errors:
            return Response({"errors": errors}, status=400)

        serializers = []
        for entry_payload in entries.values():
            serializer = self.get_serializer(data=entry_payload)
            if not serializer.is_valid():
                errors.append(
                    {
                        "entry_number": entry_payload.get("entry_number"),
                        "message": serializer.errors,
                    }
                )
            serializers.append(serializer)

        if errors:
            return Response({"errors": errors}, status=400)

        created_entries = []
        with transaction.atomic():
            for serializer in serializers:
                created_entries.append(serializer.save(created_by=request.user))

        return Response(
            {
                "created_count": len(created_entries),
                "entry_numbers": [entry.entry_number for entry in created_entries],
            }
        )


class InvoiceViewSet(AuditLogMixin, RowLevelScopeMixin, viewsets.ModelViewSet):
    queryset = (
        Invoice.objects.prefetch_related("items", "payments")
        .select_related("project", "cost_code", "customer", "created_by")
    )
    serializer_class = InvoiceSerializer
    permission_classes = [ActionBasedRolePermission]
    user_scope_fields = ("created_by", "project__created_by", "submitted_by", "approved_by", "rejected_by")
    action_role_map = {
        "list": FINANCE_READ_ROLES,
        "retrieve": FINANCE_READ_ROLES,
        "create": FINANCE_WRITE_ROLES,
        "update": FINANCE_WRITE_ROLES,
        "partial_update": FINANCE_WRITE_ROLES,
        "destroy": {ROLE_ADMIN, ROLE_ACCOUNTANT},
        "submit": FINANCE_WRITE_ROLES,
        "approve": FINANCE_APPROVER_ROLES,
        "reject": FINANCE_APPROVER_ROLES,
    }
    filterset_fields = ["invoice_type", "status", "project", "cost_code", "customer", "issue_date", "partner_name"]
    search_fields = ["invoice_number", "partner_name", "project__code", "cost_code__code"]
    ordering_fields = ["issue_date", "created_at", "invoice_number", "total_amount"]

    def get_queryset(self):
        return self.apply_row_level_scope(super().get_queryset())

    def perform_create(self, serializer):
        instance = serializer.save(created_by=self.request.user)
        self.log_action(action="create", instance=instance, changes=serializer.validated_data)

    @action(detail=True, methods=["post"])
    def submit(self, request, pk=None):
        invoice = self.get_object()
        _assert_project_open(invoice.project)
        if invoice.status != Invoice.InvoiceStatus.DRAFT:
            raise ValidationError({"status": "Only draft invoices can be submitted."})
        if not invoice.items.exists():
            raise ValidationError({"items": "Invoice must contain at least one item before submission."})
        if invoice.total_amount <= Decimal("0.00"):
            raise ValidationError({"total_amount": "Invoice total must be greater than zero."})

        invoice.status = Invoice.InvoiceStatus.PENDING_APPROVAL
        invoice.submitted_at = timezone.now()
        invoice.submitted_by = request.user
        invoice.approved_at = None
        invoice.approved_by = None
        invoice.rejected_at = None
        invoice.rejected_by = None
        invoice.rejection_reason = ""
        invoice.save(
            update_fields=[
                "status",
                "submitted_at",
                "submitted_by",
                "approved_at",
                "approved_by",
                "rejected_at",
                "rejected_by",
                "rejection_reason",
                "updated_at",
            ]
        )
        return Response(self.get_serializer(invoice).data)

    @action(detail=True, methods=["post"])
    def approve(self, request, pk=None):
        invoice = self.get_object()
        _assert_project_open(invoice.project)
        if not _is_finance_approver(request.user):
            raise PermissionDenied("You are not allowed to approve invoices.")
        if invoice.status != Invoice.InvoiceStatus.PENDING_APPROVAL:
            raise ValidationError({"status": "Only submitted invoices can be approved."})
        if not invoice.submitted_at:
            raise ValidationError({"status": "Invoice must be submitted before approval."})

        invoice.status = Invoice.InvoiceStatus.ISSUED
        invoice.approved_at = timezone.now()
        invoice.approved_by = request.user
        invoice.rejected_at = None
        invoice.rejected_by = None
        invoice.rejection_reason = ""
        invoice.save(
            update_fields=[
                "status",
                "approved_at",
                "approved_by",
                "rejected_at",
                "rejected_by",
                "rejection_reason",
                "updated_at",
            ]
        )

        if invoice.invoice_type in {Invoice.InvoiceType.SUPPLIER, Invoice.InvoiceType.SUBCONTRACTOR}:
            sync_project_cost_records_by_source(
                project=invoice.project,
                record_type=ProjectCostRecord.RecordType.ACTUAL,
                source_module="finance.invoice",
                source_reference=invoice.invoice_number,
                record_date=invoice.issue_date,
                created_by=request.user,
                amount_by_cost_code=_build_invoice_cost_amount_map(invoice),
                notes_prefix="Auto-synced from invoice ",
            )
        else:
            settle_project_cost_records_by_source(
                record_type=ProjectCostRecord.RecordType.ACTUAL,
                source_module="finance.invoice",
                source_reference=invoice.invoice_number,
            )
        _auto_post_operational_event(
            source_module="finance.invoice",
            source_event="approved",
            source_object=invoice,
            entry_date=invoice.issue_date,
            description=f"Auto-posting for approved invoice {invoice.invoice_number}",
            posted_by=request.user,
            idempotency_key=f"finance.invoice:{invoice.id}:approved",
        )
        return Response(self.get_serializer(invoice).data)

    @action(detail=True, methods=["post"])
    def reject(self, request, pk=None):
        invoice = self.get_object()
        if not _is_finance_approver(request.user):
            raise PermissionDenied("You are not allowed to reject invoices.")
        if invoice.status != Invoice.InvoiceStatus.PENDING_APPROVAL:
            raise ValidationError({"status": "Only submitted invoices can be rejected."})
        if not invoice.submitted_at:
            raise ValidationError({"status": "Invoice must be submitted before rejection."})

        reason = str(request.data.get("reason", "")).strip()
        if not reason:
            raise ValidationError({"reason": "A rejection reason is required."})

        invoice.status = Invoice.InvoiceStatus.REJECTED
        invoice.approved_at = None
        invoice.approved_by = None
        invoice.rejected_at = timezone.now()
        invoice.rejected_by = request.user
        invoice.rejection_reason = reason
        invoice.save(
            update_fields=[
                "status",
                "approved_at",
                "approved_by",
                "rejected_at",
                "rejected_by",
                "rejection_reason",
                "updated_at",
            ]
        )
        settle_project_cost_records_by_source(
            record_type=ProjectCostRecord.RecordType.ACTUAL,
            source_module="finance.invoice",
            source_reference=invoice.invoice_number,
        )
        return Response(self.get_serializer(invoice).data)


class ProgressBillingViewSet(AuditLogMixin, RowLevelScopeMixin, viewsets.ModelViewSet):
    queryset = ProgressBilling.objects.select_related(
        "project",
        "linked_invoice",
        "created_by",
        "submitted_by",
        "approved_by",
        "rejected_by",
    )
    serializer_class = ProgressBillingSerializer
    permission_classes = [ActionBasedRolePermission]
    user_scope_fields = ("project__created_by", "created_by", "submitted_by", "approved_by", "rejected_by")
    action_role_map = {
        "list": FINANCE_READ_ROLES,
        "retrieve": FINANCE_READ_ROLES,
        "create": FINANCE_WRITE_ROLES,
        "update": FINANCE_WRITE_ROLES,
        "partial_update": FINANCE_WRITE_ROLES,
        "destroy": {ROLE_ADMIN, ROLE_ACCOUNTANT},
        "submit": FINANCE_WRITE_ROLES,
        "approve": FINANCE_APPROVER_ROLES,
        "reject": FINANCE_APPROVER_ROLES,
        "generate_invoice": FINANCE_WRITE_ROLES,
    }
    filterset_fields = ["status", "project", "billing_date", "linked_invoice"]
    search_fields = ["billing_number", "project__code", "project__name"]
    ordering_fields = ["billing_date", "created_at", "billing_number", "total_amount"]

    def get_queryset(self):
        return self.apply_row_level_scope(super().get_queryset())

    def perform_create(self, serializer):
        instance = serializer.save(created_by=self.request.user)
        self.log_action(action="create", instance=instance, changes=serializer.validated_data)

    def perform_destroy(self, instance):
        if instance.status != ProgressBilling.Status.DRAFT:
            raise ValidationError({"status": "Only draft progress billings can be deleted."})
        _assert_project_open(instance.project)
        self.log_action(action="delete", instance=instance)
        instance.delete()

    @action(detail=True, methods=["post"])
    def submit(self, request, pk=None):
        billing = self.get_object()
        _assert_project_open(billing.project)
        if billing.status != ProgressBilling.Status.DRAFT:
            raise ValidationError({"status": "Only draft progress billings can be submitted."})
        if billing.completion_percentage <= Decimal("0.00"):
            raise ValidationError({"completion_percentage": "Completion percentage must be greater than zero."})

        billing.status = ProgressBilling.Status.PENDING_APPROVAL
        billing.submitted_at = timezone.now()
        billing.submitted_by = request.user
        billing.approved_at = None
        billing.approved_by = None
        billing.rejected_at = None
        billing.rejected_by = None
        billing.rejection_reason = ""
        billing.save(
            update_fields=[
                "status",
                "submitted_at",
                "submitted_by",
                "approved_at",
                "approved_by",
                "rejected_at",
                "rejected_by",
                "rejection_reason",
                "updated_at",
            ]
        )
        return Response(self.get_serializer(billing).data)

    @action(detail=True, methods=["post"])
    def approve(self, request, pk=None):
        billing = self.get_object()
        _assert_project_open(billing.project)
        if not _is_finance_approver(request.user):
            raise PermissionDenied("You are not allowed to approve progress billings.")
        if billing.status != ProgressBilling.Status.PENDING_APPROVAL:
            raise ValidationError({"status": "Only submitted progress billings can be approved."})
        if not billing.submitted_at:
            raise ValidationError({"status": "Progress billing must be submitted before approval."})

        with transaction.atomic():
            billing = (
                ProgressBilling.objects.select_for_update()
                .select_related("project")
                .get(id=billing.id)
            )
            if billing.status != ProgressBilling.Status.PENDING_APPROVAL:
                raise ValidationError({"status": "Only submitted progress billings can be approved."})

            _assert_project_open(billing.project)
            contract_value_snapshot = billing.project.contract_value
            subtotal = _quantize_money(
                contract_value_snapshot * (billing.completion_percentage / Decimal("100.00"))
            )
            if subtotal <= Decimal("0.00"):
                raise ValidationError({"completion_percentage": "Calculated amount must be greater than zero."})

            approved_subtotal = _approved_progress_billing_subtotal(billing.project, exclude_id=billing.id)
            if approved_subtotal + subtotal > contract_value_snapshot:
                raise ValidationError(
                    {
                        "completion_percentage": "Approved billings exceed project contract value after this approval."
                    }
                )

            tax_amount = _quantize_money(subtotal * (billing.tax_rate / Decimal("100.00")))
            total_amount = subtotal + tax_amount

            billing.status = ProgressBilling.Status.APPROVED
            billing.contract_value_snapshot = contract_value_snapshot
            billing.subtotal = subtotal
            billing.tax_amount = tax_amount
            billing.total_amount = total_amount
            billing.approved_at = timezone.now()
            billing.approved_by = request.user
            billing.rejected_at = None
            billing.rejected_by = None
            billing.rejection_reason = ""
            billing.save(
                update_fields=[
                    "status",
                    "contract_value_snapshot",
                    "subtotal",
                    "tax_amount",
                    "total_amount",
                    "approved_at",
                    "approved_by",
                    "rejected_at",
                    "rejected_by",
                    "rejection_reason",
                    "updated_at",
                ]
            )

        _auto_post_operational_event(
            source_module="finance.progress_billing",
            source_event="approved",
            source_object=billing,
            entry_date=billing.billing_date,
            description=f"Auto-posting for approved progress billing {billing.billing_number}",
            posted_by=request.user,
            idempotency_key=f"finance.progress_billing:{billing.id}:approved",
        )
        billing.refresh_from_db()
        return Response(self.get_serializer(billing).data)

    @action(detail=True, methods=["post"])
    def reject(self, request, pk=None):
        billing = self.get_object()
        if not _is_finance_approver(request.user):
            raise PermissionDenied("You are not allowed to reject progress billings.")
        if billing.status != ProgressBilling.Status.PENDING_APPROVAL:
            raise ValidationError({"status": "Only submitted progress billings can be rejected."})
        if not billing.submitted_at:
            raise ValidationError({"status": "Progress billing must be submitted before rejection."})

        reason = str(request.data.get("reason", "")).strip()
        if not reason:
            raise ValidationError({"reason": "A rejection reason is required."})

        billing.status = ProgressBilling.Status.REJECTED
        billing.approved_at = None
        billing.approved_by = None
        billing.rejected_at = timezone.now()
        billing.rejected_by = request.user
        billing.rejection_reason = reason
        billing.save(
            update_fields=[
                "status",
                "approved_at",
                "approved_by",
                "rejected_at",
                "rejected_by",
                "rejection_reason",
                "updated_at",
            ]
        )
        return Response(self.get_serializer(billing).data)

    @action(detail=True, methods=["post"], url_path="generate-invoice")
    def generate_invoice(self, request, pk=None):
        billing = self.get_object()
        _assert_project_open(billing.project)
        if billing.status not in {ProgressBilling.Status.APPROVED, ProgressBilling.Status.INVOICED}:
            raise ValidationError({"status": "Only approved progress billings can generate an invoice."})

        with transaction.atomic():
            billing = (
                ProgressBilling.objects.select_for_update()
                .select_related("project", "linked_invoice")
                .get(id=billing.id)
            )
            if billing.linked_invoice:
                if billing.status != ProgressBilling.Status.INVOICED:
                    billing.status = ProgressBilling.Status.INVOICED
                    billing.save(update_fields=["status", "updated_at"])
                return Response(self.get_serializer(billing).data)

            if billing.subtotal <= Decimal("0.00"):
                raise ValidationError({"subtotal": "Progress billing amount must be approved before invoicing."})

            invoice = Invoice.objects.create(
                invoice_number=next_invoice_number(),
                invoice_type=Invoice.InvoiceType.CUSTOMER,
                status=Invoice.InvoiceStatus.DRAFT,
                project=billing.project,
                partner_name=billing.project.client_name,
                issue_date=billing.billing_date,
                due_date=billing.billing_date,
                currency=billing.project.currency,
                subtotal=billing.subtotal,
                tax_amount=billing.tax_amount,
                total_amount=billing.total_amount,
                notes=f"Auto-generated from progress billing {billing.billing_number}",
                created_by=request.user,
            )
            InvoiceItem.objects.create(
                invoice=invoice,
                description=f"Progress billing {billing.billing_number}",
                quantity=Decimal("1.000"),
                unit_price=billing.subtotal,
                tax_rate=billing.tax_rate,
            )

            billing.linked_invoice = invoice
            billing.status = ProgressBilling.Status.INVOICED
            billing.save(update_fields=["linked_invoice", "status", "updated_at"])

        _auto_post_operational_event(
            source_module="finance.progress_billing",
            source_event="generate_invoice",
            source_object=billing.linked_invoice,
            entry_date=billing.billing_date,
            description=f"Auto-posting for invoice generated from progress billing {billing.billing_number}",
            posted_by=request.user,
            idempotency_key=f"finance.progress_billing:{billing.id}:generate_invoice",
        )
        billing.refresh_from_db()
        return Response(self.get_serializer(billing).data)


class RevenueRecognitionEntryViewSet(AuditLogMixin, RowLevelScopeMixin, viewsets.ModelViewSet):
    queryset = RevenueRecognitionEntry.objects.select_related(
        "project",
        "progress_billing",
        "created_by",
        "submitted_by",
        "approved_by",
        "rejected_by",
    )
    serializer_class = RevenueRecognitionEntrySerializer
    permission_classes = [ActionBasedRolePermission]
    user_scope_fields = ("project__created_by", "created_by", "submitted_by", "approved_by", "rejected_by")
    action_role_map = {
        "list": FINANCE_READ_ROLES,
        "retrieve": FINANCE_READ_ROLES,
        "create": FINANCE_WRITE_ROLES,
        "update": FINANCE_WRITE_ROLES,
        "partial_update": FINANCE_WRITE_ROLES,
        "destroy": {ROLE_ADMIN, ROLE_ACCOUNTANT},
        "submit": FINANCE_WRITE_ROLES,
        "approve": FINANCE_APPROVER_ROLES,
        "reject": FINANCE_APPROVER_ROLES,
    }
    filterset_fields = ["method", "status", "project", "recognition_date", "progress_billing"]
    search_fields = ["entry_number", "project__code", "project__name"]
    ordering_fields = ["recognition_date", "created_at", "entry_number", "recognized_amount"]

    def get_queryset(self):
        return self.apply_row_level_scope(super().get_queryset())

    def perform_create(self, serializer):
        instance = serializer.save(created_by=self.request.user)
        self.log_action(action="create", instance=instance, changes=serializer.validated_data)

    def perform_destroy(self, instance):
        if instance.status != RevenueRecognitionEntry.Status.DRAFT:
            raise ValidationError({"status": "Only draft revenue entries can be deleted."})
        self.log_action(action="delete", instance=instance)
        instance.delete()

    @action(detail=True, methods=["post"])
    def submit(self, request, pk=None):
        entry = self.get_object()
        if entry.status != RevenueRecognitionEntry.Status.DRAFT:
            raise ValidationError({"status": "Only draft revenue entries can be submitted."})

        if entry.method == RevenueRecognitionEntry.RecognitionMethod.PERCENTAGE_OF_COMPLETION:
            _assert_project_open(entry.project)
            if not entry.progress_billing and entry.recognized_percentage <= Decimal("0.00"):
                raise ValidationError(
                    {"recognized_percentage": "Provide a positive percentage or select a progress billing."}
                )
        elif entry.method == RevenueRecognitionEntry.RecognitionMethod.COMPLETED_CONTRACT:
            if entry.project.status != Project.Status.COMPLETED:
                raise ValidationError({"project": "Completed contract method requires a completed project."})

        entry.status = RevenueRecognitionEntry.Status.PENDING_APPROVAL
        entry.submitted_at = timezone.now()
        entry.submitted_by = request.user
        entry.approved_at = None
        entry.approved_by = None
        entry.rejected_at = None
        entry.rejected_by = None
        entry.rejection_reason = ""
        entry.save(
            update_fields=[
                "status",
                "submitted_at",
                "submitted_by",
                "approved_at",
                "approved_by",
                "rejected_at",
                "rejected_by",
                "rejection_reason",
                "updated_at",
            ]
        )
        return Response(self.get_serializer(entry).data)

    @action(detail=True, methods=["post"])
    def approve(self, request, pk=None):
        entry = self.get_object()
        if not _is_finance_approver(request.user):
            raise PermissionDenied("You are not allowed to approve revenue recognition entries.")
        if entry.status != RevenueRecognitionEntry.Status.PENDING_APPROVAL:
            raise ValidationError({"status": "Only submitted revenue entries can be approved."})
        if not entry.submitted_at:
            raise ValidationError({"status": "Revenue entry must be submitted before approval."})

        with transaction.atomic():
            entry = (
                RevenueRecognitionEntry.objects.select_for_update()
                .select_related("project", "progress_billing")
                .get(id=entry.id)
            )
            if entry.status != RevenueRecognitionEntry.Status.PENDING_APPROVAL:
                raise ValidationError({"status": "Only submitted revenue entries can be approved."})

            approved_total = _approved_revenue_total(entry.project, exclude_id=entry.id)
            recognized_percentage = entry.recognized_percentage

            if entry.method == RevenueRecognitionEntry.RecognitionMethod.PERCENTAGE_OF_COMPLETION:
                _assert_project_open(entry.project)
                if entry.progress_billing:
                    if entry.progress_billing.project_id != entry.project_id:
                        raise ValidationError(
                            {"progress_billing": "Progress billing must belong to the selected project."}
                        )
                    if entry.progress_billing.status not in {
                        ProgressBilling.Status.APPROVED,
                        ProgressBilling.Status.INVOICED,
                    }:
                        raise ValidationError(
                            {"progress_billing": "Progress billing must be approved before recognition."}
                        )
                    recognized_amount = entry.progress_billing.subtotal
                    if entry.project.contract_value > Decimal("0.00"):
                        recognized_percentage = _quantize_money(
                            (recognized_amount / entry.project.contract_value) * Decimal("100.00")
                        )
                else:
                    recognized_amount = _quantize_money(
                        entry.project.contract_value * (entry.recognized_percentage / Decimal("100.00"))
                    )
            elif entry.method == RevenueRecognitionEntry.RecognitionMethod.COMPLETED_CONTRACT:
                if entry.project.status != Project.Status.COMPLETED:
                    raise ValidationError({"project": "Completed contract method requires a completed project."})
                recognized_amount = _quantize_money(entry.project.contract_value - approved_total)
                recognized_percentage = Decimal("100.00")
            else:
                raise ValidationError({"method": "Unsupported revenue recognition method."})

            if recognized_amount <= Decimal("0.00"):
                raise ValidationError({"recognized_amount": "Recognized amount must be greater than zero."})

            if approved_total + recognized_amount > entry.project.contract_value:
                raise ValidationError({"recognized_amount": "Approved recognized revenue exceeds contract value."})

            entry.status = RevenueRecognitionEntry.Status.APPROVED
            entry.recognized_amount = recognized_amount
            entry.recognized_percentage = recognized_percentage
            entry.approved_at = timezone.now()
            entry.approved_by = request.user
            entry.rejected_at = None
            entry.rejected_by = None
            entry.rejection_reason = ""
            entry.save(
                update_fields=[
                    "status",
                    "recognized_amount",
                    "recognized_percentage",
                    "approved_at",
                    "approved_by",
                    "rejected_at",
                    "rejected_by",
                    "rejection_reason",
                    "updated_at",
                ]
            )

        _auto_post_operational_event(
            source_module="finance.revenue_recognition",
            source_event="approved",
            source_object=entry,
            entry_date=entry.recognition_date,
            description=f"Auto-posting for approved revenue recognition {entry.entry_number}",
            posted_by=request.user,
            idempotency_key=f"finance.revenue_recognition:{entry.id}:approved",
        )
        entry.refresh_from_db()
        return Response(self.get_serializer(entry).data)

    @action(detail=True, methods=["post"])
    def reject(self, request, pk=None):
        entry = self.get_object()
        if not _is_finance_approver(request.user):
            raise PermissionDenied("You are not allowed to reject revenue recognition entries.")
        if entry.status != RevenueRecognitionEntry.Status.PENDING_APPROVAL:
            raise ValidationError({"status": "Only submitted revenue entries can be rejected."})
        if not entry.submitted_at:
            raise ValidationError({"status": "Revenue entry must be submitted before rejection."})

        reason = str(request.data.get("reason", "")).strip()
        if not reason:
            raise ValidationError({"reason": "A rejection reason is required."})

        entry.status = RevenueRecognitionEntry.Status.REJECTED
        entry.approved_at = None
        entry.approved_by = None
        entry.rejected_at = timezone.now()
        entry.rejected_by = request.user
        entry.rejection_reason = reason
        entry.save(
            update_fields=[
                "status",
                "approved_at",
                "approved_by",
                "rejected_at",
                "rejected_by",
                "rejection_reason",
                "updated_at",
            ]
        )
        return Response(self.get_serializer(entry).data)


class PaymentViewSet(AuditLogMixin, RowLevelScopeMixin, viewsets.ModelViewSet):
    queryset = Payment.objects.select_related("invoice", "recorded_by", "submitted_by", "approved_by", "rejected_by")
    serializer_class = PaymentSerializer
    permission_classes = [ActionBasedRolePermission]
    user_scope_fields = (
        "recorded_by",
        "submitted_by",
        "approved_by",
        "rejected_by",
        "invoice__created_by",
        "invoice__project__created_by",
    )
    action_role_map = {
        "list": FINANCE_READ_ROLES,
        "retrieve": FINANCE_READ_ROLES,
        "create": FINANCE_WRITE_ROLES,
        "update": FINANCE_WRITE_ROLES,
        "partial_update": FINANCE_WRITE_ROLES,
        "destroy": {ROLE_ADMIN, ROLE_ACCOUNTANT},
        "submit": FINANCE_WRITE_ROLES,
        "approve": FINANCE_APPROVER_ROLES,
        "reject": FINANCE_APPROVER_ROLES,
    }
    filterset_fields = ["status", "method", "payment_date", "invoice"]
    search_fields = ["reference_no", "invoice__invoice_number"]
    ordering_fields = ["payment_date", "amount", "created_at"]

    def get_queryset(self):
        return self.apply_row_level_scope(super().get_queryset())

    def perform_create(self, serializer):
        invoice = serializer.validated_data["invoice"]
        _assert_project_open(invoice.project)
        instance = serializer.save(recorded_by=self.request.user)
        self.log_action(action="create", instance=instance, changes=serializer.validated_data)

    @action(detail=True, methods=["post"])
    def submit(self, request, pk=None):
        payment = self.get_object()
        _assert_project_open(payment.invoice.project)
        if payment.status != Payment.Status.PENDING:
            raise ValidationError({"status": "Only pending payments can be submitted."})
        if payment.invoice.status not in {Invoice.InvoiceStatus.ISSUED, Invoice.InvoiceStatus.PARTIALLY_PAID}:
            raise ValidationError({"invoice": "Payment can be submitted only for issued invoices."})

        payment.submitted_at = timezone.now()
        payment.submitted_by = request.user
        payment.approved_at = None
        payment.approved_by = None
        payment.rejected_at = None
        payment.rejected_by = None
        payment.rejection_reason = ""
        payment.save(
            update_fields=[
                "submitted_at",
                "submitted_by",
                "approved_at",
                "approved_by",
                "rejected_at",
                "rejected_by",
                "rejection_reason",
                "updated_at",
            ]
        )
        return Response(self.get_serializer(payment).data)

    @action(detail=True, methods=["post"])
    def approve(self, request, pk=None):
        payment = self.get_object()
        _assert_project_open(payment.invoice.project)
        if not _is_finance_approver(request.user):
            raise PermissionDenied("You are not allowed to approve payments.")
        if payment.status != Payment.Status.PENDING:
            raise ValidationError({"status": "Only pending payments can be approved."})
        if not payment.submitted_at:
            raise ValidationError({"status": "Payment must be submitted before approval."})
        if payment.invoice.status not in {Invoice.InvoiceStatus.ISSUED, Invoice.InvoiceStatus.PARTIALLY_PAID}:
            raise ValidationError({"invoice": "Payment can be approved only for issued invoices."})

        payment.status = Payment.Status.CONFIRMED
        payment.approved_at = timezone.now()
        payment.approved_by = request.user
        payment.rejected_at = None
        payment.rejected_by = None
        payment.rejection_reason = ""
        payment.save(
            update_fields=[
                "status",
                "approved_at",
                "approved_by",
                "rejected_at",
                "rejected_by",
                "rejection_reason",
                "updated_at",
            ]
        )
        _sync_invoice_payment_status(payment.invoice)
        _auto_post_operational_event(
            source_module="finance.payment",
            source_event="approved",
            source_object=payment,
            entry_date=payment.payment_date,
            description=f"Auto-posting for approved payment {payment.id}",
            posted_by=request.user,
            idempotency_key=f"finance.payment:{payment.id}:approved",
        )
        return Response(self.get_serializer(payment).data)

    @action(detail=True, methods=["post"])
    def reject(self, request, pk=None):
        payment = self.get_object()
        _assert_project_open(payment.invoice.project)
        if not _is_finance_approver(request.user):
            raise PermissionDenied("You are not allowed to reject payments.")
        if payment.status != Payment.Status.PENDING:
            raise ValidationError({"status": "Only pending payments can be rejected."})
        if not payment.submitted_at:
            raise ValidationError({"status": "Payment must be submitted before rejection."})

        reason = str(request.data.get("reason", "")).strip()
        if not reason:
            raise ValidationError({"reason": "A rejection reason is required."})

        payment.status = Payment.Status.FAILED
        payment.approved_at = None
        payment.approved_by = None
        payment.rejected_at = timezone.now()
        payment.rejected_by = request.user
        payment.rejection_reason = reason
        payment.save(
            update_fields=[
                "status",
                "approved_at",
                "approved_by",
                "rejected_at",
                "rejected_by",
                "rejection_reason",
                "updated_at",
            ]
        )
        return Response(self.get_serializer(payment).data)


class FiscalPeriodViewSet(AuditLogMixin, viewsets.ModelViewSet):
    queryset = FiscalPeriod.objects.select_related("soft_closed_by", "hard_closed_by").all()
    serializer_class = FiscalPeriodSerializer
    permission_classes = [ActionBasedRolePermission]
    action_role_map = {
        "list": FINANCE_READ_ROLES,
        "retrieve": FINANCE_READ_ROLES,
        "create": FINANCE_SETUP_ROLES,
        "update": FINANCE_SETUP_ROLES,
        "partial_update": FINANCE_SETUP_ROLES,
        "destroy": {ROLE_ADMIN},
        "soft_close": FINANCE_SETUP_ROLES,
        "hard_close": FINANCE_SETUP_ROLES,
    }
    filterset_fields = ["year", "month", "status"]
    search_fields = ["year", "month"]
    ordering_fields = ["year", "month", "start_date", "end_date"]

    @action(detail=True, methods=["post"], url_path="soft-close")
    def soft_close(self, request, pk=None):
        period = self.get_object()
        if period.status == FiscalPeriod.Status.HARD_CLOSED:
            raise ValidationError({"status": "Period is already hard-closed."})
        if period.status == FiscalPeriod.Status.SOFT_CLOSED:
            return Response(self.get_serializer(period).data)

        period.status = FiscalPeriod.Status.SOFT_CLOSED
        period.soft_closed_at = timezone.now()
        period.soft_closed_by = request.user
        period.save(update_fields=["status", "soft_closed_at", "soft_closed_by", "updated_at"])
        return Response(self.get_serializer(period).data)

    @action(detail=True, methods=["post"], url_path="hard-close")
    def hard_close(self, request, pk=None):
        period = self.get_object()
        if period.status == FiscalPeriod.Status.HARD_CLOSED:
            return Response(self.get_serializer(period).data)

        period.status = FiscalPeriod.Status.HARD_CLOSED
        if not period.soft_closed_at:
            period.soft_closed_at = timezone.now()
            period.soft_closed_by = request.user
        period.hard_closed_at = timezone.now()
        period.hard_closed_by = request.user
        period.save(
            update_fields=[
                "status",
                "soft_closed_at",
                "soft_closed_by",
                "hard_closed_at",
                "hard_closed_by",
                "updated_at",
            ]
        )
        return Response(self.get_serializer(period).data)


class ExchangeRateViewSet(AuditLogMixin, viewsets.ModelViewSet):
    queryset = ExchangeRate.objects.select_related("created_by").all()
    serializer_class = ExchangeRateSerializer
    permission_classes = [ActionBasedRolePermission]
    action_role_map = {
        "list": FINANCE_READ_ROLES,
        "retrieve": FINANCE_READ_ROLES,
        "create": FINANCE_SETUP_ROLES,
        "update": FINANCE_SETUP_ROLES,
        "partial_update": FINANCE_SETUP_ROLES,
        "destroy": {ROLE_ADMIN, ROLE_ACCOUNTANT},
    }
    filterset_fields = ["from_currency", "to_currency", "rate_date"]
    search_fields = ["from_currency", "to_currency"]
    ordering_fields = ["rate_date", "created_at", "from_currency", "to_currency"]

    def perform_create(self, serializer):
        instance = serializer.save(created_by=self.request.user)
        self.log_action(action="create", instance=instance, changes=serializer.validated_data)


class PrintSettingsViewSet(viewsets.ViewSet):
    permission_classes = [ActionBasedRolePermission]
    action_role_map = {
        "list": FINANCE_READ_ROLES,
        "retrieve": FINANCE_READ_ROLES,
        "update": FINANCE_SETUP_ROLES,
        "partial_update": FINANCE_SETUP_ROLES,
    }

    def _get_instance(self) -> PrintSettings:
        return get_print_settings()

    def list(self, request):
        instance = self._get_instance()
        return Response(PrintSettingsSerializer(instance).data)

    def retrieve(self, request, pk=None):
        instance = self._get_instance()
        return Response(PrintSettingsSerializer(instance).data)

    def update(self, request, pk=None):
        instance = self._get_instance()
        serializer = PrintSettingsSerializer(instance, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data)

    def partial_update(self, request, pk=None):
        return self.update(request, pk=pk)


class CustomerInvoiceViewSet(RowLevelScopeMixin, mixins.ListModelMixin, mixins.RetrieveModelMixin, viewsets.GenericViewSet):
    queryset = Invoice.objects.prefetch_related("items", "payments").select_related("customer", "project")
    serializer_class = InvoiceSerializer
    permission_classes = [IsAuthenticated]
    user_scope_fields = ("customer__user",)
    global_scope_role_slugs = ()
    filterset_fields = ["status", "issue_date", "due_date"]
    search_fields = ["invoice_number", "partner_name"]
    ordering_fields = ["issue_date", "created_at", "invoice_number"]

    def get_queryset(self):
        return self.apply_row_level_scope(super().get_queryset())

    @action(detail=True, methods=["get"])
    def payments(self, request, pk=None):
        invoice = self.get_object()
        payments = Payment.objects.select_related("invoice").filter(invoice=invoice).order_by("-payment_date")
        serializer = PaymentSerializer(payments, many=True)
        return Response(serializer.data)


class CustomerPaymentViewSet(RowLevelScopeMixin, mixins.ListModelMixin, mixins.RetrieveModelMixin, viewsets.GenericViewSet):
    queryset = Payment.objects.select_related("invoice").all()
    serializer_class = PaymentSerializer
    permission_classes = [IsAuthenticated]
    user_scope_fields = ("invoice__customer__user",)
    global_scope_role_slugs = ()
    filterset_fields = ["status", "payment_date", "invoice"]
    search_fields = ["reference_no", "invoice__invoice_number"]
    ordering_fields = ["payment_date", "created_at", "amount"]

    def get_queryset(self):
        return self.apply_row_level_scope(super().get_queryset())

    @action(detail=True, methods=["get"])
    def allocations(self, request, pk=None):
        payment = self.get_object()
        allocations = (
            PaymentAllocation.objects.select_related("invoice", "installment", "payment")
            .filter(payment=payment)
            .order_by("-created_at")
        )
        serializer = PaymentAllocationDetailSerializer(allocations, many=True)
        return Response(serializer.data)


class PostingRuleViewSet(AuditLogMixin, viewsets.ModelViewSet):
    queryset = PostingRule.objects.prefetch_related("lines").select_related("created_by")
    serializer_class = PostingRuleSerializer
    permission_classes = [ActionBasedRolePermission]
    action_role_map = {
        "list": FINANCE_READ_ROLES,
        "retrieve": FINANCE_READ_ROLES,
        "create": FINANCE_SETUP_ROLES,
        "update": FINANCE_SETUP_ROLES,
        "partial_update": FINANCE_SETUP_ROLES,
        "destroy": {ROLE_ADMIN, ROLE_ACCOUNTANT},
    }
    filterset_fields = ["source_module", "source_event", "is_active", "entry_class", "posting_policy"]
    search_fields = ["name", "source_module", "source_event", "description"]
    ordering_fields = ["name", "source_module", "source_event", "created_at"]

    def perform_create(self, serializer):
        instance = serializer.save(created_by=self.request.user)
        self.log_action(action="create", instance=instance, changes=serializer.validated_data)


class RecurringEntryTemplateViewSet(AuditLogMixin, viewsets.ModelViewSet):
    queryset = RecurringEntryTemplate.objects.prefetch_related("lines").select_related("created_by", "project")
    serializer_class = RecurringEntryTemplateSerializer
    permission_classes = [ActionBasedRolePermission]
    action_role_map = {
        "list": FINANCE_READ_ROLES,
        "retrieve": FINANCE_READ_ROLES,
        "create": FINANCE_SETUP_ROLES,
        "update": FINANCE_SETUP_ROLES,
        "partial_update": FINANCE_SETUP_ROLES,
        "destroy": {ROLE_ADMIN, ROLE_ACCOUNTANT},
    }
    filterset_fields = ["frequency", "is_active", "auto_post", "project", "next_run_date"]
    search_fields = ["template_code", "name", "description"]
    ordering_fields = ["template_code", "next_run_date", "created_at"]

    def perform_create(self, serializer):
        instance = serializer.save(created_by=self.request.user)
        self.log_action(action="create", instance=instance, changes=serializer.validated_data)


class BankAccountViewSet(AuditLogMixin, viewsets.ModelViewSet):
    queryset = BankAccount.objects.select_related("gl_account").all()
    serializer_class = BankAccountSerializer
    permission_classes = [ActionBasedRolePermission]
    action_role_map = {
        "list": FINANCE_READ_ROLES,
        "retrieve": FINANCE_READ_ROLES,
        "create": FINANCE_SETUP_ROLES,
        "update": FINANCE_SETUP_ROLES,
        "partial_update": FINANCE_SETUP_ROLES,
        "destroy": {ROLE_ADMIN, ROLE_ACCOUNTANT},
    }
    filterset_fields = ["currency", "is_active", "gl_account"]
    search_fields = ["code", "name", "bank_name", "account_number"]
    ordering_fields = ["code", "name", "created_at"]


class BankStatementViewSet(AuditLogMixin, viewsets.ModelViewSet):
    queryset = BankStatement.objects.prefetch_related("lines").select_related("bank_account", "uploaded_by")
    serializer_class = BankStatementSerializer
    permission_classes = [ActionBasedRolePermission]
    action_role_map = {
        "list": FINANCE_READ_ROLES,
        "retrieve": FINANCE_READ_ROLES,
        "create": FINANCE_SETUP_ROLES,
        "update": FINANCE_SETUP_ROLES,
        "partial_update": FINANCE_SETUP_ROLES,
        "destroy": {ROLE_ADMIN, ROLE_ACCOUNTANT},
    }
    filterset_fields = ["bank_account", "statement_date", "status"]
    search_fields = ["bank_account__code", "bank_account__name", "notes"]
    ordering_fields = ["statement_date", "created_at"]

    def perform_create(self, serializer):
        instance = serializer.save(uploaded_by=self.request.user)
        self.log_action(action="create", instance=instance, changes=serializer.validated_data)


class BankReconciliationSessionViewSet(AuditLogMixin, viewsets.ModelViewSet):
    queryset = BankReconciliationSession.objects.select_related(
        "bank_account", "period", "started_by", "closed_by"
    ).all()
    serializer_class = BankReconciliationSessionSerializer
    permission_classes = [ActionBasedRolePermission]
    action_role_map = {
        "list": FINANCE_READ_ROLES,
        "retrieve": FINANCE_READ_ROLES,
        "create": FINANCE_SETUP_ROLES,
        "update": FINANCE_SETUP_ROLES,
        "partial_update": FINANCE_SETUP_ROLES,
        "destroy": {ROLE_ADMIN, ROLE_ACCOUNTANT},
    }
    filterset_fields = ["bank_account", "period", "status"]
    search_fields = ["bank_account__code", "notes"]
    ordering_fields = ["started_at", "created_at"]

    def perform_create(self, serializer):
        instance = serializer.save(started_by=self.request.user)
        self.log_action(action="create", instance=instance, changes=serializer.validated_data)


class FinanceReportsViewSet(viewsets.ViewSet):
    permission_classes = [ActionBasedRolePermission]
    action_role_map = {
        "trial_balance": FINANCE_READ_ROLES,
        "general_journal": FINANCE_READ_ROLES,
        "general_ledger": FINANCE_READ_ROLES,
        "balance_sheet": FINANCE_READ_ROLES,
        "income_statement": FINANCE_READ_ROLES,
    }

    def _resolve_start_end_dates(self, request):
        start_date = _parse_date_param(request.query_params.get("start_date"), field_name="start_date")
        end_date = _parse_date_param(request.query_params.get("end_date"), field_name="end_date")
        if not end_date:
            end_date = timezone.localdate()
        if not start_date:
            start_date = end_date.replace(day=1)
        if start_date > end_date:
            raise ValidationError({"start_date": "start_date must be before or equal to end_date."})
        return start_date, end_date

    def _resolve_project_id(self, request):
        raw_project_id = request.query_params.get("project")
        if not raw_project_id:
            return None
        try:
            return int(raw_project_id)
        except (TypeError, ValueError):
            raise ValidationError({"project": "project must be an integer id."})

    @action(detail=False, methods=["get"], url_path="trial-balance")
    def trial_balance(self, request):
        start_date, end_date = self._resolve_start_end_dates(request)
        project_id = self._resolve_project_id(request)
        return Response(build_trial_balance(start_date=start_date, end_date=end_date, project_id=project_id))

    @action(detail=False, methods=["get"], url_path="general-journal")
    def general_journal(self, request):
        start_date, end_date = self._resolve_start_end_dates(request)
        project_id = self._resolve_project_id(request)
        return Response(build_general_journal(start_date=start_date, end_date=end_date, project_id=project_id))

    @action(detail=False, methods=["get"], url_path="general-ledger")
    def general_ledger(self, request):
        start_date, end_date = self._resolve_start_end_dates(request)
        project_id = self._resolve_project_id(request)
        return Response(build_general_ledger(start_date=start_date, end_date=end_date, project_id=project_id))

    @action(detail=False, methods=["get"], url_path="balance-sheet")
    def balance_sheet(self, request):
        as_of_date = _parse_date_param(request.query_params.get("as_of_date"), field_name="as_of_date")
        if not as_of_date:
            as_of_date = timezone.localdate()
        project_id = self._resolve_project_id(request)
        return Response(build_balance_sheet(as_of_date=as_of_date, project_id=project_id))

    @action(detail=False, methods=["get"], url_path="income-statement")
    def income_statement(self, request):
        start_date, end_date = self._resolve_start_end_dates(request)
        project_id = self._resolve_project_id(request)
        return Response(build_income_statement(start_date=start_date, end_date=end_date, project_id=project_id))


class YearCloseViewSet(viewsets.ViewSet):
    permission_classes = [ActionBasedRolePermission]
    action_role_map = {
        "run": FINANCE_SETUP_ROLES,
    }

    @action(detail=False, methods=["post"], url_path=r"(?P<year>\d{4})/run")
    def run(self, request, year=None):
        try:
            fiscal_year = int(year)
        except (TypeError, ValueError):
            raise ValidationError({"year": "Year must be a 4-digit number."})

        retained_earnings_account = (
            Account.objects.filter(code="3300").first()
            or Account.objects.filter(account_type=Account.AccountType.EQUITY).order_by("code").first()
        )
        if not retained_earnings_account:
            raise ValidationError(
                {"account": "Retained earnings account not found. Configure equity account code 3300 or any equity account."}
            )

        year_lines = JournalLine.objects.filter(
            entry__status=JournalEntry.Status.POSTED,
            entry__entry_date__year=fiscal_year,
            account__account_type__in=[Account.AccountType.REVENUE, Account.AccountType.EXPENSE],
        )
        aggregated = year_lines.values("account_id", "account__account_type").annotate(
            debit=Sum("debit"),
            credit=Sum("credit"),
        )
        if not aggregated:
            return Response({"detail": f"No temporary balances found for fiscal year {fiscal_year}."})

        entry_lines = []
        total_debit = Decimal("0.00")
        total_credit = Decimal("0.00")
        net_income = Decimal("0.00")

        for row in aggregated:
            account_id = row["account_id"]
            account_type = row["account__account_type"]
            debit = row["debit"] or Decimal("0.00")
            credit = row["credit"] or Decimal("0.00")

            if account_type == Account.AccountType.REVENUE:
                balance = credit - debit
                if balance == Decimal("0.00"):
                    continue
                entry_lines.append(
                    {
                        "account_id": account_id,
                        "description": f"Close revenue account {account_id}",
                        "debit": balance if balance > 0 else Decimal("0.00"),
                        "credit": -balance if balance < 0 else Decimal("0.00"),
                    }
                )
                net_income += balance
            else:
                balance = debit - credit
                if balance == Decimal("0.00"):
                    continue
                entry_lines.append(
                    {
                        "account_id": account_id,
                        "description": f"Close expense account {account_id}",
                        "debit": Decimal("0.00") if balance > 0 else -balance,
                        "credit": balance if balance > 0 else Decimal("0.00"),
                    }
                )
                net_income -= balance

        if net_income >= Decimal("0.00"):
            retained_debit = Decimal("0.00")
            retained_credit = net_income
        else:
            retained_debit = -net_income
            retained_credit = Decimal("0.00")

        entry_lines.append(
            {
                "account_id": retained_earnings_account.id,
                "description": f"Transfer net result {fiscal_year}",
                "debit": retained_debit,
                "credit": retained_credit,
            }
        )

        for line in entry_lines:
            total_debit += line["debit"]
            total_credit += line["credit"]
        if total_debit != total_credit:
            raise ValidationError(
                {"year_close": f"Year close entry is not balanced ({total_debit} debit vs {total_credit} credit)."}
            )

        closing_entry = JournalEntry.objects.create(
            entry_number=_next_related_entry_number("CLS", str(fiscal_year)),
            entry_date=date(fiscal_year, 12, 31),
            description=f"Year closing entry for {fiscal_year}",
            status=JournalEntry.Status.POSTED,
            entry_class=JournalEntry.EntryClass.CLOSING,
            currency="KWD",
            fx_rate_to_base=Decimal("1.00000000"),
            period=FiscalPeriod.objects.filter(year=fiscal_year, month=12).first(),
            posted_at=timezone.now(),
            posted_by=request.user,
            created_by=request.user,
        )
        for line in entry_lines:
            JournalLine.objects.create(
                entry=closing_entry,
                account_id=line["account_id"],
                description=line["description"],
                debit=line["debit"],
                credit=line["credit"],
            )

        return Response(
            {
                "entry_id": closing_entry.id,
                "entry_number": closing_entry.entry_number,
                "fiscal_year": fiscal_year,
                "net_income_transferred": net_income,
            }
        )
